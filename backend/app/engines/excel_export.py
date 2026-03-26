"""
Excel export engine using openpyxl.
Produces formatted, multi-sheet workbooks for each model type.
Each workbook includes:
  - An "Assumptions" tab listing all input parameters
  - Financial schedule tabs where derived rows use Excel formulas
    (e.g. Gross Profit = Revenue + COGS, EBITDA = GP + SG&A, etc.)
  - Cross-sheet references link Cash Flow to Income Statement, etc.
"""
from __future__ import annotations
import io
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
C_DARK   = "0D1117"
C_MID    = "161B22"
C_ACCENT = "1F6FEB"
C_POS    = "1A7F37"
C_NEG    = "CF222E"
C_WARN   = "9A6700"
C_LIGHT  = "F0F6FC"
C_BORDER = "30363D"

FMT_USD   = '"$"#,##0'
FMT_USD1  = '"$"#,##0.0'
FMT_PCT   = '0.0%'
FMT_X     = '0.00"x"'
FMT_RATIO = '0.00'
FMT_INT   = '#,##0'


# ── Style helpers ──────────────────────────────────────────────────────────────

def _side() -> Side:
    return Side(style="thin", color=C_BORDER)

def _border() -> Border:
    return Border(bottom=_side())

def _hf(size: int = 9, bold: bool = True) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=C_LIGHT)

def _bf(size: int = 9, bold: bool = False, color: str = C_LIGHT) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _bg(row: int) -> str:
    return "1C2128" if row % 2 == 0 else C_DARK

def _col(col: int) -> str:
    return get_column_letter(col)


def _write_header_row(ws, row: int, labels: list[str], widths: list[int] | None = None) -> None:
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = _hf()
        cell.fill = _fill(C_DARK)
        cell.alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center")
        cell.border = _border()
        if widths and c <= len(widths):
            ws.column_dimensions[_col(c)].width = widths[c - 1]


def _title_block(ws, title: str, subtitle: str) -> None:
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 14
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name="Calibri", size=13, bold=True, color=C_ACCENT)
    t.fill = _fill(C_DARK)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name="Calibri", size=9, color="8B949E")
    s.fill = _fill(C_DARK)


def _freeze(ws, at: str = "B4") -> None:
    ws.freeze_panes = at
    ws.sheet_view.showGridLines = False
    ws.tab_color = C_ACCENT


# ── Row writers ────────────────────────────────────────────────────────────────

def _write_data_row(
    ws,
    row: int,
    label: str,
    values: list[Any],
    fmt: str = FMT_USD,
    bold: bool = False,
    color: str | None = None,
) -> None:
    """Write one labelled data row. Values may be numbers OR Excel formula strings."""
    bg = _bg(row)
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _bf(bold=bold)
    lc.fill = _fill(bg)
    lc.alignment = Alignment(horizontal="left", indent=1)
    lc.border = _border()

    for col, val in enumerate(values, 2):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="right")
        c.border = _border()
        c.number_format = fmt

        if isinstance(val, str) and val.startswith("="):
            # Formula cell — colour like a bold/derived row
            c.font = _bf(bold=bold, color=color or (C_ACCENT if bold else C_LIGHT))
        elif isinstance(val, (int, float)):
            if val < 0:
                c.font = _bf(bold=bold, color=C_NEG)
            elif bold:
                c.font = _bf(bold=True, color=C_ACCENT if not color else color)
            else:
                c.font = _bf(color=C_LIGHT)
        else:
            # None / null — show dash
            c.value = "—" if val is None else val
            c.font = _bf(color="8B949E")


def _write_kv_row(ws, row: int, label: str, value: Any, fmt: str) -> None:
    """Write a single label-value pair (used in summary sheets)."""
    bg = _bg(row)
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _bf(color=C_LIGHT)
    lc.fill = _fill(bg)
    lc.border = _border()
    vc = ws.cell(row=row, column=2, value=value)
    vc.number_format = fmt
    vc.font = _bf(bold=True, color=C_ACCENT)
    vc.fill = _fill(bg)
    vc.border = _border()
    vc.alignment = Alignment(horizontal="right")


def _write_assumptions(
    ws,
    title: str,
    subtitle: str,
    params: list[tuple[str, Any, str]],
) -> None:
    """Write a labelled assumptions/inputs sheet."""
    _title_block(ws, title, subtitle)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    _write_header_row(ws, 3, ["Parameter", "Value"], [36, 22])
    for i, (label, value, fmt) in enumerate(params, 4):
        _write_kv_row(ws, i, label, value, fmt)
    ws.sheet_view.showGridLines = False
    ws.tab_color = C_WARN


# ── Formula helpers ────────────────────────────────────────────────────────────

def _frow(template: str, cols: range) -> list[str]:
    """Expand a formula template across all year columns.
    Use {col} as placeholder for the column letter, e.g. "={col}4+{col}5".
    """
    return [template.format(col=_col(c)) for c in cols]


def _xref(sheet: str, row: int, col: int) -> str:
    """Cross-sheet cell reference, e.g. ='Income Statement'!B8"""
    return f"='{sheet}'!{_col(col)}{row}"


def _xref_row(sheet: str, row: int, cols: range) -> list[str]:
    """Cross-sheet reference for each year column."""
    return [_xref(sheet, row, c) for c in cols]


# ── Corporate Export ───────────────────────────────────────────────────────────

def export_corporate(result: dict, inp: dict | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    years: list[int] = result["years"]
    n = len(years)
    year_labels = [str(y) for y in years]
    cols = range(2, n + 2)

    is_ = result["income_statement"]
    cf  = result["cash_flow"]
    tax = result["tax"]
    ds  = result["debt_schedule"]
    val = result["valuation"]

    # ── Assumptions ───────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("Assumptions")
    tl = inp.get("timeline", {}) if inp else {}
    rv = inp.get("revenue", {}) if inp else {}
    co = inp.get("costs", {}) if inp else {}
    cx = inp.get("capex", {}) if inp else {}
    wc = inp.get("working_capital", {}) if inp else {}
    db = inp.get("debt", {}) if inp else {}
    tx = inp.get("tax", {}) if inp else {}

    ass_params: list[tuple[str, Any, str]] = [
        ("Model Name",               result.get("name", ""),      "@"),
        ("─── Timeline ───",         "",                           "@"),
        ("Start Year",               tl.get("start_year", years[0] if years else ""), FMT_INT),
        ("Forecast Years",           n,                            FMT_INT),
        ("Discount Rate (WACC)",     tl.get("discount_rate", ""), FMT_PCT),
        ("Terminal Growth Rate",     tl.get("terminal_growth_rate", ""), FMT_PCT),
        ("─── Revenue ───",          "",                           "@"),
        ("Base Revenue",             rv.get("base_revenue", ""),   FMT_USD),
        ("─── Costs ───",            "",                           "@"),
        ("COGS %",                   co.get("cogs_pct", ""),       FMT_PCT),
        ("SG&A %",                   co.get("sga_pct", ""),        FMT_PCT),
        ("─── CapEx ───",            "",                           "@"),
        ("Maintenance CapEx %",      cx.get("maintenance_capex_pct", ""), FMT_PCT),
        ("Depreciation Life (yrs)",  cx.get("depreciation_years", ""),   FMT_INT),
        ("─── Working Capital ───",  "",                           "@"),
        ("DSO (days)",               wc.get("dso_days", ""),       FMT_INT),
        ("DIO (days)",               wc.get("dio_days", ""),       FMT_INT),
        ("DPO (days)",               wc.get("dpo_days", ""),       FMT_INT),
        ("─── Debt ───",             "",                           "@"),
        ("Initial Debt",             db.get("initial_debt", ""),   FMT_USD),
        ("Interest Rate",            db.get("interest_rate", ""),  FMT_PCT),
        ("Amortization (yrs)",       db.get("amortization_years", ""), FMT_INT),
        ("─── Tax ───",              "",                           "@"),
        ("Tax Rate",                 tx.get("tax_rate", ""),       FMT_PCT),
        ("NOL Carryforward",         tx.get("nol_carryforward", 0), FMT_USD),
        ("─── Key Outputs ───",      "",                           "@"),
        ("Enterprise Value",         val["enterprise_value"],      FMT_USD),
        ("Equity Value",             val["equity_value"],          FMT_USD),
        ("Net Debt",                 val["net_debt"],              FMT_USD),
        ("Equity IRR",               result.get("equity_irr") or 0, FMT_PCT),
        ("Terminal Value %",         val["terminal_value_pct"] / 100, FMT_PCT),
    ]
    _write_assumptions(ws_a, result.get("name", "Corporate Model"),
                       "Input Assumptions & Key Outputs", ass_params)

    # ── Income Statement (with formulas) ───────────────────────────────────────
    ws1 = wb.create_sheet("Income Statement")
    _title_block(ws1, result.get("name", "Corporate Model"), "Income Statement")
    _write_header_row(ws1, 3, [""] + year_labels, [28] + [12] * n)

    # Row index map for cross-sheet references
    IS = {}
    r = 4
    IS["revenue"] = r
    _write_data_row(ws1, r, "Revenue", is_["revenue"], FMT_USD, True); r += 1

    IS["cogs"] = r
    _write_data_row(ws1, r, "  COGS", [-v for v in is_["cogs"]], FMT_USD); r += 1

    IS["gross_profit"] = r
    _write_data_row(ws1, r, "Gross Profit",
                    _frow(f"={{col}}{IS['revenue']}+{{col}}{IS['cogs']}", cols),
                    FMT_USD, True); r += 1

    IS["sga"] = r
    _write_data_row(ws1, r, "  SG&A", [-v for v in is_["sga"]], FMT_USD); r += 1

    IS["ebitda"] = r
    _write_data_row(ws1, r, "EBITDA",
                    _frow(f"={{col}}{IS['gross_profit']}+{{col}}{IS['sga']}", cols),
                    FMT_USD, True); r += 1

    IS["ebitda_margin"] = r
    _write_data_row(ws1, r, "EBITDA Margin",
                    _frow(f"=IF({{col}}{IS['revenue']}<>0,{{col}}{IS['ebitda']}/{{col}}{IS['revenue']},0)", cols),
                    FMT_PCT); r += 1

    IS["depreciation"] = r
    _write_data_row(ws1, r, "  Depreciation", [-v for v in is_["depreciation"]], FMT_USD); r += 1

    IS["ebit"] = r
    _write_data_row(ws1, r, "EBIT",
                    _frow(f"={{col}}{IS['ebitda']}+{{col}}{IS['depreciation']}", cols),
                    FMT_USD, True); r += 1

    _freeze(ws1)

    # ── Cash Flow (formulas + cross-sheet refs) ────────────────────────────────
    ws2 = wb.create_sheet("Cash Flow")
    _title_block(ws2, result.get("name", "Corporate Model"), "Free Cash Flow")
    _write_header_row(ws2, 3, [""] + year_labels, [28] + [12] * n)

    CF = {}
    r = 4
    CF["ebitda"] = r
    _write_data_row(ws2, r, "EBITDA",
                    _xref_row("Income Statement", IS["ebitda"], cols),
                    FMT_USD, True); r += 1

    CF["taxes"] = r
    _write_data_row(ws2, r, "  Cash Taxes", [-v for v in tax["cash_taxes"]], FMT_USD); r += 1

    CF["capex"] = r
    _write_data_row(ws2, r, "  CapEx", [-v for v in cf["capex"]], FMT_USD); r += 1

    CF["dnwc"] = r
    _write_data_row(ws2, r, "  Δ Working Capital", [-v for v in cf["delta_nwc"]], FMT_USD); r += 1

    CF["fcf"] = r
    _write_data_row(ws2, r, "Free Cash Flow",
                    _frow(f"={{col}}{CF['ebitda']}+{{col}}{CF['taxes']}+{{col}}{CF['capex']}+{{col}}{CF['dnwc']}", cols),
                    FMT_USD, True); r += 1

    CF["interest"] = r
    _write_data_row(ws2, r, "  Interest", [-v for v in ds["interest"]], FMT_USD); r += 1

    CF["principal"] = r
    _write_data_row(ws2, r, "  Principal", [-v for v in ds["principal"]], FMT_USD); r += 1

    CF["equity_fcf"] = r
    _write_data_row(ws2, r, "Equity FCF",
                    _frow(f"={{col}}{CF['fcf']}+{{col}}{CF['interest']}+{{col}}{CF['principal']}", cols),
                    FMT_USD, True); r += 1

    _freeze(ws2)

    # ── Debt Schedule (formulas) ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Debt Schedule")
    _title_block(ws3, result.get("name", "Corporate Model"), "Debt Schedule & DSCR")
    _write_header_row(ws3, 3, [""] + year_labels, [28] + [12] * n)

    DS = {}
    r = 4
    DS["opening"] = r
    _write_data_row(ws3, r, "Opening Balance", ds["opening_balance"], FMT_USD); r += 1

    DS["new_debt"] = r
    _write_data_row(ws3, r, "New Debt", ds["new_debt"], FMT_USD); r += 1

    DS["interest"] = r
    _write_data_row(ws3, r, "  Interest", [-v for v in ds["interest"]], FMT_USD); r += 1

    DS["principal"] = r
    _write_data_row(ws3, r, "  Principal", [-v for v in ds["principal"]], FMT_USD); r += 1

    DS["closing"] = r
    _write_data_row(ws3, r, "Closing Balance",
                    _frow(f"={{col}}{DS['opening']}+{{col}}{DS['new_debt']}+{{col}}{DS['interest']}+{{col}}{DS['principal']}", cols),
                    FMT_USD, True); r += 1

    DS["fcf"] = r
    _write_data_row(ws3, r, "FCF (from Cash Flow)",
                    _xref_row("Cash Flow", CF["fcf"], cols), FMT_USD); r += 1

    DS["dscr"] = r
    # DSCR = FCF / (|Interest| + |Principal|); both stored as negative → -(int+prin)
    _write_data_row(ws3, r, "DSCR",
                    _frow(f"=IFERROR({{col}}{DS['fcf']}/(-{{col}}{DS['interest']}-{{col}}{DS['principal']}),\"—\")", cols),
                    FMT_RATIO, True); r += 1

    DS["equity_fcf"] = r
    _write_data_row(ws3, r, "Equity FCF",
                    _xref_row("Cash Flow", CF["equity_fcf"], cols), FMT_USD, True); r += 1

    _freeze(ws3)

    # ── Tax Detail ─────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Tax Detail")
    _title_block(ws4, result.get("name", "Corporate Model"), "Tax Schedule with NOL Carryforward")
    _write_header_row(ws4, 3, [""] + year_labels, [28] + [12] * n)

    TX = {}
    r = 4
    TX["ebit"] = r
    _write_data_row(ws4, r, "EBIT",
                    _xref_row("Income Statement", IS["ebit"], cols), FMT_USD, True); r += 1

    TX["nol_gen"] = r
    _write_data_row(ws4, r, "NOL Generated", tax["nol_generated"], FMT_USD); r += 1

    TX["nol_used"] = r
    _write_data_row(ws4, r, "NOL Used", [-v for v in tax["nol_used"]], FMT_USD); r += 1

    _write_data_row(ws4, r, "NOL Balance", tax["nol_balance"], FMT_USD); r += 1

    TX["taxable"] = r
    _write_data_row(ws4, r, "Taxable Income", tax["taxable_income"], FMT_USD, True); r += 1

    _write_data_row(ws4, r, "Cash Taxes", [-v for v in tax["cash_taxes"]], FMT_USD, True); r += 1
    _write_data_row(ws4, r, "Book Taxes", [-v for v in tax["book_taxes"]], FMT_USD); r += 1

    TX["eff_rate"] = r
    # Effective tax rate = Cash Taxes / max(Taxable Income, 1) — show formula
    _write_data_row(ws4, r, "Effective Tax Rate", tax["effective_tax_rate"], FMT_PCT); r += 1

    _freeze(ws4)

    # ── Valuation Summary ──────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Valuation")
    _title_block(ws5, result.get("name", "Corporate Model"), "DCF Valuation Summary")
    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 22
    ws5.sheet_view.showGridLines = False
    ws5.tab_color = C_POS

    _write_header_row(ws5, 3, ["Metric", "Value"], [32, 22])
    r = 4
    pv_start = r
    for i, pv in enumerate(val["pv_fcfs"]):
        _write_kv_row(ws5, r, f"  PV FCF — {years[i]}", pv, FMT_USD); r += 1

    pv_end = r - 1
    sum_row = r
    # Sum PV FCFs as formula
    vc = ws5.cell(row=r, column=2, value=f"=SUM(B{pv_start}:B{pv_end})")
    vc.number_format = FMT_USD
    vc.font = _bf(bold=True, color=C_ACCENT)
    vc.fill = _fill(_bg(r))
    vc.border = _border()
    vc.alignment = Alignment(horizontal="right")
    lc = ws5.cell(row=r, column=1, value="Sum PV FCFs")
    lc.font = _bf(bold=True, color=C_LIGHT)
    lc.fill = _fill(_bg(r))
    lc.border = _border()
    r += 1

    tv_row = r
    _write_kv_row(ws5, r, "PV Terminal Value", val["pv_terminal_value"], FMT_USD); r += 1

    ev_row = r
    vc = ws5.cell(row=r, column=2, value=f"=B{sum_row}+B{tv_row}")
    vc.number_format = FMT_USD
    vc.font = _bf(bold=True, color=C_ACCENT)
    vc.fill = _fill(_bg(r))
    vc.border = _border()
    vc.alignment = Alignment(horizontal="right")
    lc = ws5.cell(row=r, column=1, value="Enterprise Value")
    lc.font = _bf(bold=True, color=C_LIGHT)
    lc.fill = _fill(_bg(r))
    lc.border = _border()
    r += 1

    nd_row = r
    _write_kv_row(ws5, r, "Net Debt", val["net_debt"], FMT_USD); r += 1

    eq_row = r
    vc = ws5.cell(row=r, column=2, value=f"=B{ev_row}-B{nd_row}")
    vc.number_format = FMT_USD
    vc.font = _bf(bold=True, color=C_ACCENT)
    vc.fill = _fill(_bg(r))
    vc.border = _border()
    vc.alignment = Alignment(horizontal="right")
    lc = ws5.cell(row=r, column=1, value="Equity Value")
    lc.font = _bf(bold=True, color=C_LIGHT)
    lc.fill = _fill(_bg(r))
    lc.border = _border()
    r += 1

    tv_pct_row = r
    vc = ws5.cell(row=r, column=2, value=f"=IF(B{ev_row}<>0,B{tv_row}/B{ev_row},0)")
    vc.number_format = FMT_PCT
    vc.font = _bf(color=C_LIGHT)
    vc.fill = _fill(_bg(r))
    vc.border = _border()
    vc.alignment = Alignment(horizontal="right")
    lc = ws5.cell(row=r, column=1, value="Terminal Value % of EV")
    lc.font = _bf(color=C_LIGHT)
    lc.fill = _fill(_bg(r))
    lc.border = _border()
    r += 1

    _write_kv_row(ws5, r, "Equity IRR", result.get("equity_irr") or 0, FMT_PCT)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Project Finance Export ────────────────────────────────────────────────────

def export_project(result: dict, inp: dict | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    years: list[int] = result["years_operating"]
    n = len(years)
    year_labels = [str(y) for y in years]
    cols = range(2, n + 2)

    op = result["operating"]
    ds = result["debt_schedule"]
    tax = result["tax"]
    val = result["valuation"]
    km  = result["key_metrics"]

    # ── Assumptions ───────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("Assumptions")
    ph = inp.get("phases", {}) if inp else {}
    rv = inp.get("revenue", {}) if inp else {}
    co = inp.get("costs", {}) if inp else {}
    db = inp.get("debt", {}) if inp else {}
    tl = inp.get("timeline", {}) if inp else {}
    tx = inp.get("tax", {}) if inp else {}

    ass_params: list[tuple[str, Any, str]] = [
        ("Model Name",               result.get("name", ""),          "@"),
        ("─── Timeline ───",         "",                               "@"),
        ("Construction Years",       ph.get("construction_years", ""), FMT_INT),
        ("Operating Years",          ph.get("operating_years", n),     FMT_INT),
        ("Discount Rate",            tl.get("discount_rate", ""),      FMT_PCT),
        ("─── Revenue ───",          "",                               "@"),
        ("Capacity (MW)",            rv.get("capacity_mw", ""),        FMT_INT),
        ("Capacity Factor",          rv.get("capacity_factor", ""),    FMT_PCT),
        ("Price per Unit ($/MWh)",   rv.get("price_per_unit", ""),     FMT_USD1),
        ("Price Escalation",         rv.get("price_escalation", ""),   FMT_PCT),
        ("─── Costs ───",            "",                               "@"),
        ("Construction Cost",        co.get("construction_cost", ""),  FMT_USD),
        ("Equity %",                 co.get("equity_pct", ""),         FMT_PCT),
        ("Annual O&M",               co.get("opex_per_year", ""),      FMT_USD),
        ("OpEx Escalation",          co.get("opex_escalation", ""),    FMT_PCT),
        ("─── Debt ───",             "",                               "@"),
        ("Debt Interest Rate",       db.get("interest_rate", ""),      FMT_PCT),
        ("Amortization (yrs)",       db.get("amortization_years", ""), FMT_INT),
        ("DSRA Months",              db.get("debt_service_reserve_months", ""), FMT_INT),
        ("Grace Period (yrs)",       db.get("grace_period_years", 0),  FMT_INT),
        ("─── Tax ───",              "",                               "@"),
        ("Tax Rate",                 tx.get("tax_rate", ""),           FMT_PCT),
        ("─── Key Outputs ───",      "",                               "@"),
        ("Construction Cost",        result["construction_cost"],       FMT_USD),
        ("Total Debt",               result["total_debt"],              FMT_USD),
        ("Equity Investment",        result["equity_investment"],       FMT_USD),
        ("Enterprise Value",         val["enterprise_value"],           FMT_USD),
        ("Project IRR",              km.get("project_irr") or 0,       FMT_PCT),
        ("Equity IRR",               km.get("equity_irr") or 0,        FMT_PCT),
        ("Min DSCR",                 km["min_dscr"] or 0,              FMT_RATIO),
        ("Avg DSCR",                 km["avg_dscr"] or 0,              FMT_RATIO),
    ]
    _write_assumptions(ws_a, result.get("name", "Project Finance"),
                       "Input Assumptions & Key Outputs", ass_params)

    # ── Operations (with formulas) ─────────────────────────────────────────────
    ws1 = wb.create_sheet("Operations")
    _title_block(ws1, result.get("name", "Project Finance"), "Operating Projections")
    _write_header_row(ws1, 3, [""] + year_labels, [28] + [12] * n)

    OP = {}
    r = 4
    OP["revenue"] = r
    _write_data_row(ws1, r, "Revenue", op["revenue"], FMT_USD, True); r += 1

    OP["opex"] = r
    _write_data_row(ws1, r, "  OpEx", [-v for v in op["opex"]], FMT_USD); r += 1

    OP["ebitda"] = r
    _write_data_row(ws1, r, "EBITDA",
                    _frow(f"={{col}}{OP['revenue']}+{{col}}{OP['opex']}", cols),
                    FMT_USD, True); r += 1

    OP["ebitda_margin"] = r
    _write_data_row(ws1, r, "EBITDA Margin",
                    _frow(f"=IF({{col}}{OP['revenue']}<>0,{{col}}{OP['ebitda']}/{{col}}{OP['revenue']},0)", cols),
                    FMT_PCT); r += 1

    OP["dep"] = r
    _write_data_row(ws1, r, "  Depreciation", [-v for v in op["depreciation"]], FMT_USD); r += 1

    OP["ebit"] = r
    _write_data_row(ws1, r, "EBIT",
                    _frow(f"={{col}}{OP['ebitda']}+{{col}}{OP['dep']}", cols),
                    FMT_USD, True); r += 1

    OP["tax"] = r
    _write_data_row(ws1, r, "  Cash Taxes", [-v for v in tax["cash_taxes"]], FMT_USD); r += 1

    OP["fcf"] = r
    _write_data_row(ws1, r, "Free Cash Flow",
                    _frow(f"={{col}}{OP['ebitda']}+{{col}}{OP['tax']}", cols),
                    FMT_USD, True); r += 1

    _freeze(ws1)

    # ── Debt Schedule (with formulas) ──────────────────────────────────────────
    ws2 = wb.create_sheet("Debt Schedule")
    _title_block(ws2, result.get("name", "Project Finance"), "Debt Schedule & DSCR")
    _write_header_row(ws2, 3, [""] + year_labels, [28] + [12] * n)

    DS = {}
    r = 4
    DS["opening"] = r
    _write_data_row(ws2, r, "Opening Balance", ds["opening_balance"], FMT_USD); r += 1

    DS["interest"] = r
    _write_data_row(ws2, r, "  Interest", [-v for v in ds["interest"]], FMT_USD); r += 1

    DS["principal"] = r
    _write_data_row(ws2, r, "  Principal", [-v for v in ds["principal"]], FMT_USD); r += 1

    DS["closing"] = r
    _write_data_row(ws2, r, "Closing Balance",
                    _frow(f"={{col}}{DS['opening']}+{{col}}{DS['interest']}+{{col}}{DS['principal']}", cols),
                    FMT_USD, True); r += 1

    DS["fcf"] = r
    _write_data_row(ws2, r, "FCF (from Operations)",
                    _xref_row("Operations", OP["fcf"], cols), FMT_USD); r += 1

    DS["dscr"] = r
    _write_data_row(ws2, r, "DSCR",
                    _frow(f"=IFERROR({{col}}{DS['fcf']}/(-{{col}}{DS['interest']}-{{col}}{DS['principal']}),\"—\")", cols),
                    FMT_RATIO, True); r += 1

    _write_data_row(ws2, r, "DSRA Balance", ds["dsra_balance"], FMT_USD); r += 1

    DS["equity_dist"] = r
    _write_data_row(ws2, r, "Equity Distributions", ds["equity_distributions"], FMT_USD, True); r += 1

    _freeze(ws2)

    # ── Summary ────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    _title_block(ws3, result.get("name", "Project Finance"), "Key Metrics")
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 22
    ws3.sheet_view.showGridLines = False
    ws3.tab_color = C_POS
    _write_header_row(ws3, 3, ["Metric", "Value"], [32, 22])

    r = 4
    for label, value, fmt in [
        ("Construction Cost",   result["construction_cost"],  FMT_USD),
        ("Total Debt",          result["total_debt"],          FMT_USD),
        ("Equity Investment",   result["equity_investment"],   FMT_USD),
        ("Enterprise Value",    val["enterprise_value"],       FMT_USD),
        ("Equity Value",        val["equity_value"],           FMT_USD),
        ("Project IRR",         km.get("project_irr") or 0,   FMT_PCT),
        ("Equity IRR",          km.get("equity_irr") or 0,    FMT_PCT),
        ("Min DSCR",            km["min_dscr"] or 0,          FMT_RATIO),
        ("Avg DSCR",            km["avg_dscr"] or 0,          FMT_RATIO),
    ]:
        _write_kv_row(ws3, r, label, value, fmt); r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Acquisition Export ────────────────────────────────────────────────────────

def export_acquisition(result: dict, inp: dict | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    years: list[int] = result["years"]
    n = len(years)
    year_labels = [str(y) for y in years]
    cols = range(2, n + 2)
    ops = result["operations"]
    entry = result["entry"]
    exit_ = result["exit"]
    ret = result["returns"]

    # ── Assumptions ───────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("Assumptions")
    ass_params: list[tuple[str, Any, str]] = [
        ("Model Name",            result.get("name", ""),           "@"),
        ("─── Deal Structure ─────────", "", "@"),
        ("Purchase Price",        inp.get("purchase_price", "") if inp else entry["purchase_price"], FMT_USD),
        ("Entry EV/EBITDA",       inp.get("entry_ebitda_multiple", "") if inp else entry["entry_multiple"], FMT_RATIO),
        ("Exit EV/EBITDA",        inp.get("exit_ebitda_multiple", "") if inp else exit_["exit_multiple"], FMT_RATIO),
        ("Holding Period (yrs)",  inp.get("holding_period_years", n) if inp else n, FMT_INT),
        ("Equity %",              inp.get("equity_pct", "") if inp else "", FMT_PCT),
        ("Debt Interest Rate",    inp.get("debt_interest_rate", "") if inp else "", FMT_PCT),
        ("EBITDA Growth Rate",    inp.get("ebitda_growth_rate", "") if inp else "", FMT_PCT),
        ("Revenue Synergies",     inp.get("revenue_synergies", 0) if inp else 0, FMT_USD),
        ("Cost Savings",          inp.get("cost_savings", 0) if inp else 0, FMT_USD),
        ("Tax Rate",              inp.get("tax_rate", "") if inp else "", FMT_PCT),
        ("─── Key Outputs ─────────", "", "@"),
        ("Entry EBITDA",          entry["entry_ebitda"],            FMT_USD),
        ("Equity Invested",       entry["equity_invested"],          FMT_USD),
        ("Entry Debt",            entry["debt"],                     FMT_USD),
        ("Exit EV",               exit_["exit_ev"],                  FMT_USD),
        ("Exit Equity Proceeds",  exit_["exit_equity_proceeds"],     FMT_USD),
        ("Equity IRR",            ret.get("equity_irr") or 0,       FMT_PCT),
        ("MOIC",                  ret["moic"],                       FMT_X),
    ]
    _write_assumptions(ws_a, result.get("name", "Acquisition"),
                       "Deal Assumptions & Returns", ass_params)

    # ── Operations (with formulas) ─────────────────────────────────────────────
    ws1 = wb.create_sheet("Operations")
    _title_block(ws1, result.get("name", "Acquisition"), "Operating Projections")
    _write_header_row(ws1, 3, [""] + year_labels, [28] + [12] * n)

    OP = {}
    r = 4
    OP["ebitda"] = r
    _write_data_row(ws1, r, "EBITDA", ops["ebitda"], FMT_USD); r += 1

    OP["synergies"] = r
    _write_data_row(ws1, r, "  Synergies", ops["synergies"], FMT_USD); r += 1

    OP["adj_ebitda"] = r
    _write_data_row(ws1, r, "Adj. EBITDA",
                    _frow(f"={{col}}{OP['ebitda']}+{{col}}{OP['synergies']}", cols),
                    FMT_USD, True); r += 1

    OP["interest"] = r
    _write_data_row(ws1, r, "  Interest", [-v for v in ops["interest"]], FMT_USD); r += 1

    OP["ebt"] = r
    _write_data_row(ws1, r, "EBT",
                    _frow(f"={{col}}{OP['adj_ebitda']}+{{col}}{OP['interest']}", cols),
                    FMT_USD, True); r += 1

    OP["taxes"] = r
    _write_data_row(ws1, r, "  Taxes", [-v for v in ops["taxes"]], FMT_USD); r += 1

    OP["net_income"] = r
    _write_data_row(ws1, r, "Net Income",
                    _frow(f"={{col}}{OP['ebt']}+{{col}}{OP['taxes']}", cols),
                    FMT_USD, True); r += 1

    OP["principal"] = r
    _write_data_row(ws1, r, "  Principal", [-v for v in ops["principal"]], FMT_USD); r += 1

    OP["fcf"] = r
    _write_data_row(ws1, r, "FCF to Equity",
                    _frow(f"={{col}}{OP['net_income']}+{{col}}{OP['principal']}", cols),
                    FMT_USD, True); r += 1

    _write_data_row(ws1, r, "Debt Balance", ops["debt_balance"], FMT_USD); r += 1

    _freeze(ws1)

    # ── Returns ────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Returns")
    _title_block(ws2, result.get("name", "Acquisition"), "Returns Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22
    ws2.sheet_view.showGridLines = False
    ws2.tab_color = C_POS
    _write_header_row(ws2, 3, ["Metric", "Value"], [32, 22])

    r = 4
    for label, value, fmt in [
        ("Purchase Price",        entry["purchase_price"],       FMT_USD),
        ("Entry EBITDA",          entry["entry_ebitda"],          FMT_USD),
        ("Entry Multiple",        entry["entry_multiple"],        FMT_RATIO),
        ("Equity Invested",       entry["equity_invested"],       FMT_USD),
        ("Entry Debt",            entry["debt"],                  FMT_USD),
        ("Exit EBITDA",           exit_["exit_ebitda"],           FMT_USD),
        ("Exit EV",               exit_["exit_ev"],               FMT_USD),
        ("Exit Multiple",         exit_["exit_multiple"],         FMT_RATIO),
        ("Exit Debt",             exit_["exit_debt"],             FMT_USD),
        ("Exit Equity Proceeds",  exit_["exit_equity_proceeds"],  FMT_USD),
        ("Equity IRR",            ret.get("equity_irr") or 0,    FMT_PCT),
        ("MOIC",                  ret["moic"],                    FMT_X),
    ]:
        _write_kv_row(ws2, r, label, value, fmt); r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Monte Carlo Export ─────────────────────────────────────────────────────────

def export_monte_carlo(result: dict, inp: dict | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Assumptions ───────────────────────────────────────────────────────────
    if inp:
        ws_a = wb.create_sheet("Assumptions")
        ass_params: list[tuple[str, Any, str]] = [
            ("─── Simulation Parameters ─────", "", "@"),
            ("Base Revenue",          inp.get("base_revenue", ""),          FMT_USD),
            ("Revenue Volatility",    inp.get("revenue_volatility", ""),    FMT_PCT),
            ("Base OpEx",             inp.get("base_opex", ""),             FMT_USD),
            ("OpEx Volatility",       inp.get("opex_volatility", ""),       FMT_PCT),
            ("Discount Rate",         inp.get("discount_rate", ""),         FMT_PCT),
            ("Terminal Growth Rate",  inp.get("terminal_growth_rate", ""),  FMT_PCT),
            ("Forecast Years",        inp.get("forecast_years", ""),        FMT_INT),
            ("N Simulations",         inp.get("n_simulations", ""),         FMT_INT),
            ("Mean Reversion Speed",  inp.get("mean_reversion_speed", ""),  FMT_RATIO),
            ("Rev / OpEx Correlation",inp.get("correlation_rev_opex", ""),  FMT_RATIO),
        ]
        _write_assumptions(ws_a, "Monte Carlo Simulation",
                           "Simulation Parameters", ass_params)

    # ── Distribution ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Distribution")
    _title_block(ws, "Monte Carlo Simulation", f"{result['n_simulations']:,} simulations")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.sheet_view.showGridLines = False
    _write_header_row(ws, 3, ["Metric", "Value"], [24, 22])

    pcts = result["percentiles"]
    r = 4
    for label, value, fmt in [
        ("Mean EV",          result["mean_ev"],   FMT_USD),
        ("Std Dev",          result["std_ev"],    FMT_USD),
        ("Min EV",           result["min_ev"],    FMT_USD),
        ("Max EV",           result["max_ev"],    FMT_USD),
        ("Prob. Positive",   result["probability_positive"], FMT_PCT),
        ("─── Percentiles ─────", "", "@"),
        ("P1",               pcts["p1"],          FMT_USD),
        ("P5",               pcts["p5"],          FMT_USD),
        ("P10",              pcts["p10"],         FMT_USD),
        ("P25",              pcts["p25"],         FMT_USD),
        ("P50 (Median)",     pcts["p50"],         FMT_USD),
        ("P75",              pcts["p75"],         FMT_USD),
        ("P90",              pcts["p90"],         FMT_USD),
        ("P95",              pcts["p95"],         FMT_USD),
        ("P99",              pcts["p99"],         FMT_USD),
    ]:
        _write_kv_row(ws, r, label, value, fmt); r += 1

    # ── Histogram Data ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Histogram Data")
    _title_block(ws2, "Histogram", "Enterprise Value frequency distribution")
    _write_header_row(ws2, 3, ["Bin Centre ($)", "Frequency", "Cumulative %"], [22, 14, 16])
    ws2.sheet_view.showGridLines = False

    hist = result["histogram"]
    total = sum(hist["counts"])
    cum = 0
    for i, (centre, count) in enumerate(zip(hist["bin_centers"], hist["counts"]), 4):
        cum += count
        ws2.cell(row=i, column=1, value=centre).number_format = FMT_USD
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=cum / total if total else 0).number_format = FMT_PCT
        for c in [1, 2, 3]:
            bg = _bg(i)
            ws2.cell(row=i, column=c).fill = _fill(bg)
            ws2.cell(row=i, column=c).font = _bf(color=C_LIGHT)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
